import os
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from sap_manager import InboundPackageInfo


class TemplatePrinter:
	def __init__(self, src):
		self.wb = load_workbook(src)
		self.ws = self.wb["Template"]
		self.destination = Path(src).parent / "temp.xlsx"

	def write_workbook(self, package_info):
		package_info: InboundPackageInfo
		self.ws.sheet_view.showGridLines = False
		self.ws["B1"] = package_info.case_number
		self.ws["B1"].font = Font(size=14, italic=True)
		
		self.ws["B3"] = package_info.zt01_number
		self.ws["B3"].font = Font(size=14, italic=True)
		
		self.ws["B5"] = package_info.vl01n_number
		self.ws["B5"].font = Font(size=14, italic=True)
		
		self.ws["F1"] = (package_info.first_name + " " + package_info.last_name)
		self.ws["F1"].font = Font(size=14, italic=True)
		
		self.ws["F3"] = package_info.box_name
		self.ws["F3"].font = Font(size=14, italic=True)
		
		self.ws["F5"] = package_info.date_received
		self.ws["F5"].font = Font(size=14, italic=True)
		
		self.ws["E10"] = package_info.notes
		self.ws["E10"].font = Font(size=14, italic=True)
		self.ws["E10"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
	
	def save_workbook(self):
		self.wb.save(self.destination)
	
	def print_workbook(self):
		os.startfile(self.destination, 'print')