import json
import subprocess
import time
from datetime import date
from tkinter import Tk
import win32com.client
import win32com.client
from pywinauto import Application
from pywinauto_recorder.player import *
import usaddress
import os
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from sap_manager import InboundPackageInfo

def get_customer_ids(item_count):
    time.sleep(5)
    dict = {}
    src = r"C:\Users\abu89\PycharmProjects\SAPAutomation\assets\AccountNumTemplate.xlsx"
    wb = load_workbook(src)
    ws = wb["Template"]
    destination = Path(src).parent / "temp2.xlsx"
    for i in range(item_count):
        with UIPath(u"Document List||Window"):
            time.sleep(2)
            send_keys("^c")
            document_number = Tk().clipboard_get()
            send_keys("{F2}")
            time.sleep(2)
            send_keys("{TAB}")
            account_number = -1
            for j in range(10):
                send_keys("^c")
                account_number = Tk().clipboard_get().split("\t")[5]
                if len(account_number) == 6:
                    break
                send_keys("{DOWN}")
            if len(account_number) != 6:
                account_number = "INVALID ACCOUNT NUMBER. CHECK ME!"
            send_keys("{F3}")
            time.sleep(1)
            send_keys("{DOWN}")
        ws[f"A{i+2}"] = document_number
        ws[f"B{i+2}"] = account_number
        wb.save(destination)
        print(f"{i+1}/{item_count}")


get_customer_ids(430)