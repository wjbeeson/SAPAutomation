from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font
from win32api import ShellExecute
from utilities import PackageInfo


class InboundLabelPrinter:
    def __init__(self):
        src = r"assets/LabelTemplate.xlsx"
        self.wb = load_workbook(src)
        self.ws = self.wb["Sheet1"]
        self.destination = Path(src).parent / "temp_label.xlsx"

    def get_model_name(self, sku):
        sku_book = load_workbook(r"assets/SKUReference.xlsx")
        sku_sheet = sku_book["Sheet1"]
        i = 1
        name_cell = ""
        while True:
            cell = "D" + str(i)
            current_sku = str(sku_sheet[cell].value)
            if current_sku == sku:
                name_cell = sku_sheet["E" + str(i)].value
            if current_sku is None or current_sku == "None":
                break
            i = i + 1
        return name_cell

    def write_workbook(self, package_info):
        package_info: PackageInfo
        model_name = self.get_model_name(package_info.product_sku)

        self.ws.sheet_view.showGridLines = False
        self.ws["A1"] = package_info.case_number
        self.ws["A1"].font = Font(name='Arial Nova Cond', size=36, italic=False)

        self.ws["A2"] = (package_info.first_name + " " + package_info.last_name)
        self.ws["A2"].font = Font(name='Arial Nova Cond', size=36, italic=False)

        self.ws["A3"] = model_name
        self.ws["A3"].font = Font(name='Arial Nova Cond', size=36, italic=False)

        self.ws["A4"] = package_info.date_received
        self.ws["A4"].font = Font(name='Arial Nova Cond', size=36, italic=False)

    def save_workbook(self):
        self.ws.sheet_properties.pageSetUpPr.fitToPage = True
        self.wb.save(self.destination)

    def print_workbook(self):
        printer_name = 'Brother QL-720NW'
        ShellExecute(0, "printto", str(self.destination), f'"{printer_name}"', ".", 0)
